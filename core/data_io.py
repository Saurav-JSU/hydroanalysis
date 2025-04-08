"""
Functions for reading and writing hydrological data.
"""

import pandas as pd
import numpy as np
import os
import logging
from openpyxl import load_workbook

from hydroanalysis.config import CONFIG

logger = logging.getLogger(__name__)

def read_discharge_data(file_path, station_id=None, sheet_name=None):
    """
    Read discharge data from Excel or CSV files.
    
    Parameters
    ----------
    file_path : str
        Path to the data file
    station_id : str or float, optional
        ID of the station to extract
    sheet_name : str, optional
        Name of the sheet for Excel files
        
    Returns
    -------
    pandas.DataFrame
        DataFrame with discharge data
    """
    logger.info(f"Reading discharge data from {file_path}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if file_ext == '.xlsx' or file_ext == '.xls':
            if sheet_name:
                if station_id is not None:
                    # Read specific station data from Excel
                    return _read_station_discharge_excel(file_path, station_id, sheet_name)
                else:
                    # Read all discharge data from sheet
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                # Try to discover the right sheet
                wb = load_workbook(file_path, read_only=True)
                possible_sheets = ['discharge', 'Discharge', 'flow', 'Flow', 
                                  'Mean_daily_discharge', 'Daily_discharge']
                
                for sheet in possible_sheets:
                    if sheet in wb.sheetnames:
                        logger.info(f"Found discharge sheet: {sheet}")
                        if station_id is not None:
                            return _read_station_discharge_excel(file_path, station_id, sheet)
                        else:
                            df = pd.read_excel(file_path, sheet_name=sheet)
                        break
                else:
                    # If no matching sheet found
                    logger.warning("No discharge sheet found. Using first sheet.")
                    if station_id is not None:
                        return _read_station_discharge_excel(file_path, station_id, wb.sheetnames[0])
                    else:
                        df = pd.read_excel(file_path, sheet_name=0)
                
        elif file_ext == '.csv':
            # Read from CSV
            df = pd.read_csv(file_path)
            
            # If station_id is specified, extract that station's data
            if station_id is not None:
                if str(station_id) in df.columns:
                    df = df[['Date', str(station_id)]].rename(columns={str(station_id): 'Discharge'})
                else:
                    logger.warning(f"Station {station_id} not found in CSV columns")
        else:
            logger.error(f"Unsupported file extension: {file_ext}")
            return None
            
    except Exception as e:
        logger.error(f"Error reading discharge data: {e}")
        return None
    
    # Ensure Date column is datetime
    if 'Date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Date']):
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    
    # Validate discharge data
    if 'Discharge' in df.columns:
        # Check for negative values
        neg_count = (df['Discharge'] < 0).sum()
        if neg_count > 0:
            logger.warning(f"Found {neg_count} negative discharge values. These will be replaced with zeros.")
            df.loc[df['Discharge'] < 0, 'Discharge'] = 0
    
    logger.info(f"Successfully read discharge data with {len(df)} rows")
    return df

def _read_station_discharge_excel(file_path, station_id, sheet_name):
    """Helper function to extract a specific station's discharge from Excel."""
    wb = load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    
    # Get headers to find station column
    headers = list(next(sheet.values))
    try:
        station_col_idx = headers.index(station_id)
    except ValueError:
        logger.error(f"Station {station_id} not found in headers")
        return None
    
    # Find date column
    date_col_idx = None
    for i, header in enumerate(headers):
        if header and isinstance(header, str) and header.lower() in ['date', 'datetime', 'time']:
            date_col_idx = i
            break
    
    if date_col_idx is None:
        # If no column is labeled date, assume it's the first or second column
        date_col_idx = 0 if len(headers) > 1 else 1
    
    # Extract date and discharge
    dates = []
    discharge = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) > max(date_col_idx, station_col_idx):
            if row[date_col_idx] is not None and row[station_col_idx] is not None:
                dates.append(row[date_col_idx])
                discharge.append(row[station_col_idx])
    
    # Create DataFrame
    df = pd.DataFrame({
        'Date': dates,
        'Discharge': discharge
    })
    
    # Ensure date is datetime
    if not pd.api.types.is_datetime64_any_dtype(df['Date']):
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    
    logger.info(f"Extracted {len(df)} discharge records for station {station_id}")
    return df

def read_precipitation_data(file_path, sheet_name=None):
    """
    Read precipitation data from Excel or CSV files.
    
    Parameters
    ----------
    file_path : str
        Path to the data file
    sheet_name : str, optional
        Name of the sheet for Excel files
        
    Returns
    -------
    pandas.DataFrame
        DataFrame with precipitation data
    """
    logger.info(f"Reading precipitation data from {file_path}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if file_ext in ['.xlsx', '.xls']:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                # Try to discover the right sheet
                wb = load_workbook(file_path, read_only=True)
                possible_sheets = ['precipitation', 'Precipitation', 'rainfall', 'Rainfall', 
                                  'Complete', 'Daily']
                
                for sheet in possible_sheets:
                    if sheet in wb.sheetnames:
                        logger.info(f"Found precipitation sheet: {sheet}")
                        df = pd.read_excel(file_path, sheet_name=sheet)
                        break
                else:
                    # If no matching sheet found
                    logger.warning("No precipitation sheet found. Using first sheet.")
                    df = pd.read_excel(file_path, sheet_name=0)
        
        elif file_ext == '.csv':
            # Read from CSV
            df = pd.read_csv(file_path)
        else:
            logger.error(f"Unsupported file extension: {file_ext}")
            return None
            
    except Exception as e:
        logger.error(f"Error reading precipitation data: {e}")
        return None
    
    # Ensure 'Date' column is datetime
    if 'Date' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['Date']):
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    
    # Ensure 'datetime' column is datetime if it exists
    if 'datetime' in df.columns and not pd.api.types.is_datetime64_any_dtype(df['datetime']):
        df['datetime'] = pd.to_datetime(df['datetime'], errors='coerce')
    
    # Data validation - basic checks
    extreme_value_threshold = CONFIG['comparison']['extreme_value_threshold']
    station_columns = [col for col in df.columns if 'Station_' in col]
    
    for station in station_columns:
        # Check for negative values
        neg_count = (df[station] < 0).sum()
        if neg_count > 0:
            logger.warning(f"Found {neg_count} negative values in {station}. Replacing with zeros.")
            df.loc[df[station] < 0, station] = 0
        
        # Check for unrealistically high values
        extreme_count = (df[station] > extreme_value_threshold).sum()
        if extreme_count > 0:
            logger.warning(f"Found {extreme_count} extreme values (>{extreme_value_threshold}mm) in {station}.")
    
    logger.info(f"Successfully read precipitation data with {len(df)} rows")
    
    return df

def read_station_metadata(file_path, sheet_name=None):
    """
    Read station metadata from Excel file.
    
    Parameters
    ----------
    file_path : str
        Path to the metadata file
    sheet_name : str, optional
        Name of the sheet for Excel files
        
    Returns
    -------
    pandas.DataFrame
        DataFrame with standardized station metadata
    """
    logger.info(f"Reading station metadata from {file_path}")
    
    try:
        # If sheet_name is specified, use it
        if sheet_name:
            metadata_df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            # Try to discover the right sheet
            wb = load_workbook(file_path, read_only=True)
            possible_sheets = ['metadata', 'Metadata', 'stations', 'Stations', 'Meta_data', 'Metadata']
            
            for sheet in possible_sheets:
                if sheet in wb.sheetnames:
                    logger.info(f"Found metadata sheet: {sheet}")
                    metadata_df = pd.read_excel(file_path, sheet_name=sheet)
                    break
            else:
                # If no matching sheet found
                logger.warning("No metadata sheet found. Using first sheet.")
                metadata_df = pd.read_excel(file_path, sheet_name=0)
        
        # Try to identify relevant columns
        station_id_col = None
        lat_col = None
        lon_col = None
        river_col = None
        location_col = None
        elev_col = None
        
        # Check for station ID column
        possible_id_cols = ['Station_ID', 'Station ID', 'StationID', 'Station', 'ID', 
                           'station_id', 'St. No.', 'station id']
        for col in possible_id_cols:
            if col in metadata_df.columns:
                station_id_col = col
                break
        
        # Check for latitude column
        possible_lat_cols = ['Latitude', 'Lat', 'latitude', 'lat', 'Y', 'y_coord']
        for col in possible_lat_cols:
            if col in metadata_df.columns:
                lat_col = col
                break
        
        # Check for longitude column
        possible_lon_cols = ['Longitude', 'Long', 'longitude', 'long', 'lon', 'X', 'x_coord']
        for col in possible_lon_cols:
            if col in metadata_df.columns:
                lon_col = col
                break
        
        # Check for river name column
        possible_river_cols = ['River', 'river', 'River_name', 'river_name', 'Stream', 'stream']
        for col in possible_river_cols:
            if col in metadata_df.columns:
                river_col = col
                break
        
        # Check for location column
        possible_location_cols = ['Location', 'location', 'Place', 'place', 'Site', 'site']
        for col in possible_location_cols:
            if col in metadata_df.columns:
                location_col = col
                break
        
        # Check for elevation column
        possible_elev_cols = ['Elevation', 'Elev', 'elevation', 'elev', 'alt', 'Alt', 'Altitude']
        for col in possible_elev_cols:
            if col in metadata_df.columns:
                elev_col = col
                break
        
        # Create standardized DataFrame with found columns
        standardized = {}
        
        if station_id_col:
            standardized['Station_ID'] = metadata_df[station_id_col]
        else:
            logger.warning("Station ID column not found in metadata")
            standardized['Station_ID'] = [f"Station_{i}" for i in range(len(metadata_df))]
        
        if lat_col:
            standardized['Latitude'] = metadata_df[lat_col]
        else:
            logger.warning("Latitude column not found in metadata")
            standardized['Latitude'] = np.nan
        
        if lon_col:
            standardized['Longitude'] = metadata_df[lon_col]
        else:
            logger.warning("Longitude column not found in metadata")
            standardized['Longitude'] = np.nan
        
        if river_col:
            standardized['River'] = metadata_df[river_col]
        
        if location_col:
            standardized['Location'] = metadata_df[location_col]
        
        if elev_col:
            standardized['Elevation'] = metadata_df[elev_col]
        
        # Add any other columns that might be useful
        for col in metadata_df.columns:
            if col not in [station_id_col, lat_col, lon_col, river_col, location_col, elev_col]:
                standardized[col] = metadata_df[col]
        
        standardized_df = pd.DataFrame(standardized)
        
        # Ensure Station_ID starts with 'Station_' if it's numeric
        if all(str(id).replace('.', '', 1).isdigit() for id in standardized_df['Station_ID']):
            standardized_df['Station_ID'] = 'Station_' + standardized_df['Station_ID'].astype(str)
        
        logger.info(f"Successfully processed metadata for {len(standardized_df)} stations")
        return standardized_df
    
    except Exception as e:
        logger.error(f"Error reading station metadata: {e}")
        return None

def save_data(df, file_path):
    """
    Save DataFrame to file (CSV or Excel).
    
    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame to save
    file_path : str
        Path to save file
    
    Returns
    -------
    bool
        True if successful, False otherwise
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if file_ext == '.xlsx':
            df.to_excel(file_path, index=False)
        elif file_ext == '.csv':
            df.to_csv(file_path, index=False)
        else:
            logger.warning(f"Unsupported file extension: {file_ext}. Saving as CSV instead.")
            file_path = os.path.splitext(file_path)[0] + '.csv'
            df.to_csv(file_path, index=False)
        
        logger.info(f"Data saved to {file_path}")
        return True
    except Exception as e:
        logger.error(f"Error saving data: {e}")
        return False